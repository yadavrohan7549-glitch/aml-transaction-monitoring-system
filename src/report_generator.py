"""
Step 9 - Automated Excel reports.

Produces the standard set of outputs an AML analyst/team lead would
expect to hand to a QA reviewer or regulator: a full investigation
workbook, a customer risk report, daily/monthly alert summaries, a
high-risk customer list, an SAR-style suspicious activity report, and
an executive dashboard summary - all as formatted .xlsx files.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import settings
from src.logger import log_event

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _autosize_and_style(writer, sheet_name, df):
    ws = writer.sheets[sheet_name]
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        max_len = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
    ws.freeze_panes = "A2"


def generate_investigation_report(alerts, cases, customers, transactions, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        alerts.to_excel(writer, sheet_name="Alerts", index=False)
        _autosize_and_style(writer, "Alerts", alerts)

        cases.to_excel(writer, sheet_name="Cases", index=False)
        _autosize_and_style(writer, "Cases", cases)

        merged = cases.merge(customers[["customer_id", "name", "country", "risk_level"]],
                              on="customer_id", how="left")
        merged.to_excel(writer, sheet_name="Case Detail", index=False)
        _autosize_and_style(writer, "Case Detail", merged)
    log_event("REPORTING", f"AML Investigation Report saved to {path}")


def generate_customer_risk_report(customers, risk_scores, path):
    merged = customers.merge(risk_scores, on="customer_id", how="left")
    cols = ["customer_id", "name", "country", "customer_type", "pep_status",
            "sanction_status", "risk_score", "risk_band", "contributing_factors", "total_alerts"]
    merged = merged[cols].sort_values("risk_score", ascending=False)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Customer Risk", index=False)
        _autosize_and_style(writer, "Customer Risk", merged)
    log_event("REPORTING", f"Customer Risk Report saved to {path}")


def generate_daily_alert_report(alerts, transactions, path):
    merged = alerts.merge(transactions[["transaction_id", "timestamp"]], on="transaction_id", how="left")
    merged["date"] = pd.to_datetime(merged["timestamp"]).dt.date
    daily = merged.groupby(["date", "rule_name"]).size().reset_index(name="alert_count")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        daily.to_excel(writer, sheet_name="Daily Alerts", index=False)
        _autosize_and_style(writer, "Daily Alerts", daily)
    log_event("REPORTING", f"Daily Alert Report saved to {path}")


def generate_monthly_summary(alerts, transactions, path):
    merged = alerts.merge(transactions[["transaction_id", "timestamp", "amount_usd"]],
                           on="transaction_id", how="left")
    merged["month"] = pd.to_datetime(merged["timestamp"]).dt.to_period("M").astype(str)
    monthly = merged.groupby("month").agg(
        alert_count=("rule_name", "count"),
        total_flagged_amount_usd=("amount_usd", "sum"),
    ).reset_index()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)
        _autosize_and_style(writer, "Monthly Summary", monthly)
    log_event("REPORTING", f"Monthly Summary saved to {path}")


def generate_high_risk_list(customers, risk_scores, path):
    merged = customers.merge(risk_scores, on="customer_id", how="left")
    high = merged[merged["risk_band"].isin(["High", "Critical"])].sort_values(
        "risk_score", ascending=False)
    cols = ["customer_id", "name", "country", "pep_status", "sanction_status",
            "risk_score", "risk_band", "contributing_factors"]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        high[cols].to_excel(writer, sheet_name="High Risk Customers", index=False)
        _autosize_and_style(writer, "High Risk Customers", high[cols])
    log_event("REPORTING", f"High Risk Customer List saved to {path}")


def generate_sar_report(cases, customers, transactions, path):
    sar_cases = cases[cases["risk_score"] >= 45].copy()
    merged = sar_cases.merge(customers[["customer_id", "name", "country", "occupation"]],
                              on="customer_id", how="left")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="SAR Candidates", index=False)
        _autosize_and_style(writer, "SAR Candidates", merged)
    log_event("REPORTING", f"Suspicious Activity Report saved to {path}")


def generate_executive_report(customers, transactions, alerts, cases, risk_scores, path):
    summary = pd.DataFrame([{
        "generated_on": datetime.now().isoformat(timespec="seconds"),
        "total_customers": len(customers),
        "total_transactions": len(transactions),
        "total_transaction_value_usd": round(transactions["amount_usd"].sum(), 2),
        "total_alerts": len(alerts),
        "total_cases": len(cases),
        "high_risk_customers": int((risk_scores["risk_band"] == "High").sum()),
        "critical_risk_customers": int((risk_scores["risk_band"] == "Critical").sum()),
        "pep_customers": int(customers["pep_status"].sum()),
        "sanctioned_customers": int(customers["sanction_status"].sum()),
    }])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Executive Summary", index=False)
        _autosize_and_style(writer, "Executive Summary", summary)
    log_event("REPORTING", f"Executive Dashboard Report saved to {path}")


def generate_all_reports(customers, transactions, merchants, alerts, cases, risk_scores):
    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    r = settings.REPORTS_DIR
    generate_investigation_report(alerts, cases, customers, transactions,
                                   os.path.join(r, "AML_Investigation_Report.xlsx"))
    generate_customer_risk_report(customers, risk_scores,
                                   os.path.join(r, "Customer_Risk_Report.xlsx"))
    generate_daily_alert_report(alerts, transactions,
                                 os.path.join(r, "Daily_Alert_Report.xlsx"))
    generate_monthly_summary(alerts, transactions,
                              os.path.join(r, "Monthly_Summary_Report.xlsx"))
    generate_high_risk_list(customers, risk_scores,
                             os.path.join(r, "High_Risk_Customer_List.xlsx"))
    generate_sar_report(cases, customers, transactions,
                         os.path.join(r, "Suspicious_Activity_Report.xlsx"))
    generate_executive_report(customers, transactions, alerts, cases, risk_scores,
                               os.path.join(r, "Executive_Dashboard_Report.xlsx"))
    log_event("REPORTING", "All 7 reports generated successfully")
